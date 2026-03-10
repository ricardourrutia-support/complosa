import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

st.set_page_config(page_title="Compensaciones y Contactabilidad", layout="wide")
st.title("📦 Compensaciones y Contactabilidad — Cabify Style")

# ------------------------------
# Funciones de Negocio y Compensación
# ------------------------------
def calcular_compensacion(minutos):
    if pd.isna(minutos):
        return 9000
    try:
        minutos = float(minutos)
    except:
        return 9000

    if minutos >= 50:
        return 9000
    elif minutos >= 40:
        return 6000
    elif minutos >= 35:
        return 3000
    else:
        return 0

# ------------------------------
# Función para arreglar fechas
# ------------------------------
MESES_ES_EN = {
    'enero': 'january', 'febrero': 'february', 'marzo': 'march',
    'abril': 'april', 'mayo': 'may', 'junio': 'june',
    'julio': 'july', 'agosto': 'august', 'septiembre': 'september',
    'octubre': 'october', 'noviembre': 'november', 'diciembre': 'december'
}

def convertir_fecha_espanol(fecha_str):
    if pd.isna(fecha_str):
        return pd.NaT
    
    s = str(fecha_str).lower().strip()
    s = s.replace(' de ', ' ')
    
    for es, en in MESES_ES_EN.items():
        if es in s:
            s = s.replace(es, en)
            break
            
    try:
        # Usamos dayfirst=True por si viene en formato DD/MM/YYYY
        return pd.to_datetime(s, dayfirst=True).date()
    except:
        return pd.NaT

# ------------------------------
# Funciones de Validación (Contactabilidad y Desempeño)
# ------------------------------
def es_email_contactable(email):
    if pd.isna(email): return False
    email = str(email).strip().lower()
    if email == "11@gmail.com": return False
    patron = r"^[\w\.-]+@[\w\.-]+\.\w{2,}$"
    if not re.match(patron, email): return False
    parte_local = email.split('@')[0]
    if len(parte_local) < 2: return False
    return True

def es_email_cumplimiento(email):
    if pd.isna(email): return False
    email = str(email).strip().lower()
    if email == "11@gmail.com": return True
    return es_email_contactable(email)

def es_telefono_contactable(telefono):
    if pd.isna(telefono): return False
    telefono = str(telefono).strip().replace("+", "").replace(" ", "")
    if not telefono.isdigit(): return False
    if telefono == "111111111": return False
    if len(telefono) < 8 or len(telefono) > 15: return False
    if len(set(telefono)) == 1: return False
    return True

def es_telefono_cumplimiento(telefono):
    if pd.isna(telefono): return False
    telefono_limpio = str(telefono).strip().replace("+", "").replace(" ", "")
    if telefono_limpio == "111111111": return True
    return es_telefono_contactable(telefono)

# ------------------------------
# Subir archivo CSV
# ------------------------------
uploaded_file = st.file_uploader("📤 Sube tu archivo CSV", type=["csv"])

if uploaded_file is not None:

    try:
        # Lectura inicial robusta
        df_raw = pd.read_csv(uploaded_file, dtype=str, sep=';', on_bad_lines='skip')
        st.success("Archivo cargado correctamente 🎉 (Las filas con errores de formato fueron omitidas)")
    except Exception as e:
        st.error(f"❌ Error al leer el CSV: {e}")
        st.stop()

    columnas = [
        "Día de tm_start_local_at",
        "Segmento Tiempo en Losa",
        "End State",
        "id_reservation_id",
        "Service Channel",
        "Minutes Creation - Pickup",
        "User Fullname",
        "User Email",
        "User Phone Number",
        "Service Agent"
    ]

    faltantes = [c for c in columnas if c not in df_raw.columns]
    if faltantes:
        st.error(f"❌ Faltan columnas requeridas en tu CSV: {faltantes}")
        st.stop()

    df = df_raw[columnas].copy()

    # Convertir la fecha en texto español a fecha real
    df["Día de tm_start_local_at"] = df["Día de tm_start_local_at"].apply(convertir_fecha_espanol)

    # ==========================================
    # ANALÍTICA DE AGENTES (SOBRE LA BASE COMPLETA SIN FILTRAR FECHAS)
    # ==========================================
    st.markdown("---")
    st.subheader("🕵️‍♂️ Analítica de Agentes: Contactabilidad vs Desempeño (Base Completa)")
    st.write("*(Esta sección analiza **todos los registros** de la base cargada, sin aplicar filtros de fecha ni montos de compensación).*")
    
    # Trabajamos con una copia para no alterar el DataFrame original
    df_agentes = df.copy()
    
    # Manejar los Service Agent nulos o vacíos para que no se pierdan en el GroupBy
    df_agentes["Service Agent"] = df_agentes["Service Agent"].fillna("Sin Agente / No Aplica")
    
    df_agentes["Email Contactable"] = df_agentes["User Email"].apply(es_email_contactable)
    df_agentes["Email Cumplimiento"] = df_agentes["User Email"].apply(es_email_cumplimiento)
    df_agentes["Teléfono Contactable"] = df_agentes["User Phone Number"].apply(es_telefono_contactable)
    df_agentes["Teléfono Cumplimiento"] = df_agentes["User Phone Number"].apply(es_telefono_cumplimiento)
    
    resumen_agentes = df_agentes.groupby("Service Agent").agg(
        Total_Casos=("id_reservation_id", "count"),
        Desempeño_Email=("Email Cumplimiento", "sum"),
        Contactables_Email=("Email Contactable", "sum"),
        Desempeño_Tel=("Teléfono Cumplimiento", "sum"),
        Contactables_Tel=("Teléfono Contactable", "sum")
    ).reset_index()
    
    resumen_agentes["% Desempeño Email"] = (resumen_agentes["Desempeño_Email"] / resumen_agentes["Total_Casos"]) * 100
    resumen_agentes["% Contactabilidad Email"] = (resumen_agentes["Contactables_Email"] / resumen_agentes["Total_Casos"]) * 100
    resumen_agentes["% Desempeño Teléfono"] = (resumen_agentes["Desempeño_Tel"] / resumen_agentes["Total_Casos"]) * 100
    resumen_agentes["% Contactabilidad Teléfono"] = (resumen_agentes["Contactables_Tel"] / resumen_agentes["Total_Casos"]) * 100
    
    # Ordenar por Total de Casos (descendente) para ver los agentes con más volumen arriba
    resumen_agentes = resumen_agentes.sort_values("Total_Casos", ascending=False)
    
    resumen_display = resumen_agentes.copy()
    for col in ["% Desempeño Email", "% Contactabilidad Email", "% Desempeño Teléfono", "% Contactabilidad Teléfono"]:
        resumen_display[col] = resumen_display[col].round(1).astype(str) + "%"
    
    st.dataframe(resumen_display, use_container_width=True)

    output_agentes = BytesIO()
    resumen_agentes.to_excel(output_agentes, index=False, engine='openpyxl')
    output_agentes.seek(0)
    
    st.download_button(
        "⬇️ Descargar Analítica de Agentes (Excel)",
        data=output_agentes,
        file_name="analitica_agentes_cabify.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.markdown("---")

    # ==========================================
    # FILTRO DE FECHAS (SÓLO APLICA PARA COMPENSACIONES)
    # ==========================================
    # Eliminar NaT de las fechas solo para la lógica de compensaciones
    df_fechas_validas = df.dropna(subset=["Día de tm_start_local_at"]).copy()
    
    if df_fechas_validas.empty:
        st.error("❌ No hay fechas válidas para continuar con el cálculo de compensaciones. Revisa el formato de la columna 'Día de tm_start_local_at'.")
        st.stop()

    fecha_min = df_fechas_validas["Día de tm_start_local_at"].min()
    fecha_max = df_fechas_validas["Día de tm_start_local_at"].max()

    fecha_desde, fecha_hasta = st.date_input(
        "📅 Selecciona rango de fechas para evaluar Compensaciones:",
        value=(fecha_min, fecha_max)
    )

    df_comp_filtrado = df_fechas_validas[
        (df_fechas_validas["Día de tm_start_local_at"] >= fecha_desde) &
        (df_fechas_validas["Día de tm_start_local_at"] <= fecha_hasta)
    ].copy()

    if df_comp_filtrado.empty:
        st.warning("⚠️ No hay registros en ese rango de fechas para procesar compensaciones.")
        st.stop()

    # ==========================================
    # SECCIÓN DE COMPENSACIONES
    # ==========================================
    df_comp_filtrado["Estado Pago"] = "No Pagado"
    df_comp_filtrado["Minutes Creation - Pickup"] = pd.to_numeric(df_comp_filtrado["Minutes Creation - Pickup"], errors="coerce")
    df_comp_filtrado["Monto a Reembolsar"] = df_comp_filtrado["Minutes Creation - Pickup"].apply(calcular_compensacion)

    df_compensaciones = df_comp_filtrado[df_comp_filtrado["Monto a Reembolsar"] > 0].copy()

    if df_compensaciones.empty:
        st.warning("⚠️ No hay compensaciones > 0 para mostrar en este rango.")
        st.stop()

    st.subheader("📊 Registros procesados para Compensación (Losa)")
    
    df_compensaciones_vista = df_compensaciones.copy()
    st.dataframe(df_compensaciones_vista, use_container_width=True)

    st.subheader("📈 Resumen de compensaciones")
    resumen = df_compensaciones["Monto a Reembolsar"].value_counts().sort_index()
    
    st.write("### 🧮 Cantidad de casos por monto:")
    for monto, cantidad in resumen.items():
        st.write(f"- **${monto:,}** → {cantidad} casos")

    st.write("### 💰 Total compensaciones:")
    st.write(f"- **Total de casos:** {len(df_compensaciones)}")
    st.write(f"- **Total en dinero:** ${df_compensaciones['Monto a Reembolsar'].sum():,}")

    # ==========================================
    # CREAR EXCEL ESTILIZADO CABIFY PARA COMPENSACIONES
    # ==========================================
    columnas_excel = [
        "Día de tm_start_local_at", 
        "Segmento Tiempo en Losa", 
        "End State", 
        "id_reservation_id", 
        "Service Channel", 
        "Minutes Creation - Pickup", 
        "User Fullname", 
        "User Email", 
        "User Phone Number", 
        "Estado Pago", 
        "Monto a Reembolsar"
    ]
    df_excel = df_compensaciones[columnas_excel].copy()
    df_excel = df_excel.rename(columns={"Día de tm_start_local_at": "Day of tm_start_local_at"})

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Compensaciones"

    header_fill = PatternFill("solid", fgColor="5B34AC")
    header_font = Font(color="FFFFFF", bold=True)
    border_side = Side(style="thin", color="362065")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    color_no_pagado = PatternFill("solid", fgColor="EA8C2E")
    color_pagado = PatternFill("solid", fgColor="0C936B")

    fill_3000 = PatternFill("solid", fgColor="EFBD03")
    fill_6000 = PatternFill("solid", fgColor="EA8C2E")
    fill_9000 = PatternFill("solid", fgColor="E83C96")
    alt_fill = PatternFill("solid", fgColor="FAF8FE")

    for r in dataframe_to_rows(df_excel, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = f"A1:{chr(64 + len(df_excel.columns))}1"

    col_idx_estado = df_excel.columns.get_loc("Estado Pago")
    col_idx_monto = df_excel.columns.get_loc("Monto a Reembolsar")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        idx = row[0].row

        if idx % 2 == 0:
            for cell in row:
                cell.fill = alt_fill

        estado = row[col_idx_estado].value
        if estado == "No Pagado":
            row[col_idx_estado].fill = color_no_pagado
        else:
            row[col_idx_estado].fill = color_pagado

        monto = float(row[col_idx_monto].value)
        if monto == 3000:
            row[col_idx_monto].fill = fill_3000
        elif monto == 6000:
            row[col_idx_monto].fill = fill_6000
        elif monto == 9000:
            row[col_idx_monto].fill = fill_9000

        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    dv = DataValidation(type="list", formula1='"Pagado,No Pagado"', allow_blank=False)
    col_estado_pago = col_idx_estado + 1
    col_letter = chr(64 + col_estado_pago)
    dv.add(f"{col_letter}2:{col_letter}1048576")
    ws.add_data_validation(dv)

    wb.save(output)
    output.seek(0)

    st.download_button(
        "⬇️ Descargar Excel Compensaciones",
        output,
        file_name="compensaciones_losa_cabify.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ==========================================
    # PLANTILLAS PARA ZENDESK
    # ==========================================
    st.markdown("---")
    st.subheader("📝 Plantillas para Zendesk")
    st.write("Copia los campos para crear los tickets. Solo se muestran usuarios con correos válidos y se excluyen dominios corporativos (@cabify.com).")
    
    # Recalcular las banderas de contacto solo para el grupo de compensaciones para usarlas en Zendesk
    df_compensaciones["Email Contactable"] = df_compensaciones["User Email"].apply(es_email_contactable)
    
    df_zendesk = df_compensaciones[
        (df_compensaciones["Email Contactable"] == True) & 
        (~df_compensaciones["User Email"].str.lower().str.contains("@cabify.com", na=False))
    ].copy()

    if df_zendesk.empty:
        st.info("No hay usuarios válidos para generar plantillas de Zendesk en este momento.")
    else:
        def obtener_primer_nombre(nombre_completo):
            if pd.isna(nombre_completo):
                return "Usuario"
            partes = str(nombre_completo).strip().split()
            if partes:
                return partes[0].capitalize()
            return "Usuario"

        df_zendesk["Primer Nombre"] = df_zendesk["User Fullname"].apply(obtener_primer_nombre)

        with st.expander("Ver plantillas de respuesta generadas", expanded=True):
            for idx, row in df_zendesk.iterrows():
                st.write(f"### 🎫 Ticket para: {row['User Fullname']} ({row['Monto a Reembolsar']} CLP)")
                
                col1, col2 = st.columns([1, 1.5])
                
                with col1:
                    st.write("**Datos del Ticket (Copia solo el valor exacto):**")
                    st.markdown("**Motivo:**")
                    st.code("Compensación por tu experiencia reciente con Cabify", language="text")
                    st.markdown("**Email Solicitante:**")
                    st.code(row['User Email'], language="text")
                    st.markdown("**Motivo de Contacto:**")
                    st.code("Tag 060134 (Retraso en Reserva)", language="text")
                    st.markdown("**Descuentos:**")
                    st.code("Chile -> Disculpas", language="text")
                    st.markdown("**Macro:**")
                    st.code("Compensación proactiva espera en losa", language="text")

                with col2:
                    st.write("**Mensaje a enviar (Selecciona con el ratón y copia para mantener los enlaces activos):**")
                    mensaje_correo = f'''Hola {row['Primer Nombre']},

En Cabify, valoramos tu tiempo y sabemos que cada minuto cuenta.

Queremos extender nuestras más sinceras disculpas porque en tu reciente viaje desde el Aeropuerto experimentaste una espera inusualmente larga para abordar tu vehículo. Entendemos la frustración y lamentamos no haber cumplido con nuestro estándar de servicio ágil y cómodo.

Para recuperar tu confianza y asegurarnos de que tu próxima experiencia con Cabify sea impecable, hemos preparado un saldo de cortesía en tu cuenta.

Si ya tienes una cuenta Cabify, respóndenos con el correo de tu cuenta para cargar el saldo.  
Si aún no tienes una cuenta, crea tu cuenta en menos de 2 minutos y envíanos el correo que registraste.

📱 [IPhone](https://apps.apple.com/es/app/cabify-viaja-como-te-mereces/id476087442)  
📱 [Android](https://play.google.com/store/apps/details?id=com.cabify.rider&referrer=adjust_reftag%3Dczy6okB7ZfVXv%26utm_source%3DLandings%2B-Bot%25C3%25B3n%2Bdescarga%2Bapp%26utm_campaign%3DALL-ALL-RIDER-NA-NA-LAND-NA-DM-NA-PASAJEROSNEWLP-GLO-NA-NA-NA&pli=1)

Una vez que nos confirmes o crees tu cuenta, cargaremos tu saldo en menos de 24 horas para que puedas usarlo en tu próximo viaje.

Estamos atentos para asistirte con la carga del saldo. ¡Esperamos verte pronto a bordo, con la comodidad y rapidez que mereces!

Saludos cordiales,  
El equipo de Cabify'''
                    st.info(mensaje_correo)
                
                st.markdown("---")

else:
    st.info("Sube un archivo CSV para comenzar.")
